import streamlit as st
from PIL import Image, ImageFilter
import io
import os
import zipfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation

APP_TITLE = "印刷向けサイズ指定＋Excel台帳（プルダウン＋自動調整付き）"
st.title(APP_TITLE)

SAVE_DIR = "resized"
os.makedirs(SAVE_DIR, exist_ok=True)

CM_PER_INCH = 2.54

def cm_to_inches(cm): return cm / CM_PER_INCH
def inches_to_cm(inch): return inch * CM_PER_INCH
def px_to_cm(px, dpi): return inches_to_cm(px / dpi)
def cm_to_px(cm, dpi): return round(cm_to_inches(cm) * dpi)
def describe_size(px_w, px_h, dpi):
    w_cm = px_to_cm(px_w, dpi)
    h_cm = px_to_cm(px_h, dpi)
    return w_cm, h_cm

# ========= 入力UI =========
prefix = st.text_input("保存ファイル名に付けるPrefix", value="resized")

mode = st.selectbox(
    "サイズ指定モード",
    [
        "① DPIを直接指定（ピクセルはそのまま）",
        "② 幅（cm）だけ指定 → DPIを自動計算（ピクセルはそのまま）",
        "③ 高さ（cm）だけ指定 → DPIを自動計算（ピクセルはそのまま）",
        "④ 幅×高さ（cm）とDPIを指定 → そのピクセル数にリサンプリング（ピクセルを変更）"
    ],
    index=0
)

col1, col2, col3 = st.columns(3)
with col1:
    dpi_input = st.number_input("DPI", min_value=50, max_value=1200, value=300)
with col2:
    width_cm_input = st.number_input("幅 (cm)", min_value=0.0, value=0.0)
with col3:
    height_cm_input = st.number_input("高さ (cm)", min_value=0.0, value=0.0)

format_choice = st.selectbox("保存形式を選んでください", ["JPEG", "PNG"])
max_width = st.number_input("出力画像の最大幅(px)", min_value=500, max_value=8000, value=1200)
preview = st.checkbox("処理後のプレビューを表示する", value=True)

uploaded_files = st.file_uploader(
    "画像をアップロードしてください",
    type=["jpg", "jpeg", "png", "tif", "tiff", "bmp", "webp"],
    accept_multiple_files=True
)

st.caption("②③はピクセル数を変えずに印刷サイズを調整。④はピクセル数も変わります。最大幅を超える場合は縮小されます。")

# ========= メイン処理 =========
if uploaded_files and prefix.strip():
    processed_items = []

    for uploaded_file in uploaded_files:
        img = Image.open(uploaded_file)
        img = img.convert("RGB") if format_choice == "JPEG" else img
        orig_w, orig_h = img.size
        out_img = img
        out_dpi = dpi_input

        # --- モードごとの処理 ---
        if mode == "① DPIを直接指定（ピクセルはそのまま）":
            out_dpi = dpi_input
        elif mode == "② 幅（cm）だけ指定 → DPIを自動計算（ピクセルはそのまま）":
            out_dpi = round(orig_w / cm_to_inches(width_cm_input)) if width_cm_input > 0 else dpi_input
        elif mode == "③ 高さ（cm）だけ指定 → DPIを自動計算（ピクセルはそのまま）":
            out_dpi = round(orig_h / cm_to_inches(height_cm_input)) if height_cm_input > 0 else dpi_input
        elif mode == "④ 幅×高さ（cm）とDPIを指定 → ピクセル数を変更":
            if width_cm_input > 0 and height_cm_input > 0:
                target_w = cm_to_px(width_cm_input, dpi_input)
                target_h = cm_to_px(height_cm_input, dpi_input)
                out_img = img.resize((target_w, target_h), Image.LANCZOS)
                if max(orig_w, orig_h) > max(target_w, target_h):
                    out_img = out_img.filter(ImageFilter.SHARPEN)
                out_dpi = dpi_input

        # --- 最大幅制御 ---
        if out_img.width > max_width:
            new_height = int(out_img.height * (max_width / out_img.width))
            out_img = out_img.resize((max_width, new_height), Image.LANCZOS)

        # --- 保存 ---
        ext = "jpg" if format_choice == "JPEG" else "png"
        new_filename = f"{prefix}_{os.path.splitext(uploaded_file.name)[0]}.{ext}"
        save_path = os.path.join(SAVE_DIR, new_filename)

        img_bytes = io.BytesIO()
        if format_choice == "JPEG":
            out_img.save(save_path, format="JPEG", quality=90, subsampling=0, dpi=(out_dpi, out_dpi))
            out_img.save(img_bytes, format="JPEG", quality=90, subsampling=0, dpi=(out_dpi, out_dpi))
        else:
            out_img.save(save_path, format="PNG", dpi=(out_dpi, out_dpi), compress_level=0)
            out_img.save(img_bytes, format="PNG", dpi=(out_dpi, out_dpi), compress_level=0)
        img_bytes.seek(0)

        out_w, out_h = out_img.size
        print_w_cm, print_h_cm = describe_size(out_w, out_h, out_dpi)
        info_text = f"元: {orig_w}×{orig_h}px → 出力: {out_w}×{out_h}px, DPI={out_dpi}（印刷時 {print_w_cm:.1f}×{print_h_cm:.1f} cm）"

        if preview:
            st.image(out_img, caption=f"{new_filename}\n{info_text}", use_column_width=True)
        else:
            st.caption(f"{new_filename} | {info_text}")

        processed_items.append((len(processed_items)+1, new_filename, save_path, info_text))

    # ---- ZIPまとめ ----
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zipf:
        for _, fname, path, _ in processed_items:
            with open(path, "rb") as f:
                zipf.writestr(fname, f.read())
    zip_buffer.seek(0)
    st.download_button("処理済み画像をまとめてダウンロード (ZIP)", zip_buffer, "processed_images.zip")

    # ---- Excel台帳作成（プルダウン＋行高さ自動調整）----
    wb = Workbook()
    ws = wb.active
    ws.title = "写真台帳"
    ws.append(["No", "写真", "ファイル名", "部材種類", "損傷種類", "損傷範囲/状態", "備考"])

    row = 2
    max_width_px = 300  # 画像の最大幅(px)

    for no, fname, path, info in processed_items:
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=3, value=fname)

        # 空欄セル（プルダウン対象）
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value="")

        # --- 縦横比を維持してリサイズ ---
        xl_img = XLImage(path)
        aspect_ratio = xl_img.height / xl_img.width
        new_width = max_width_px
        new_height = int(new_width * aspect_ratio)

        xl_img.width, xl_img.height = new_width, new_height
        ws.add_image(xl_img, f"B{row}")

        # --- 行高と列幅を画像に合わせて調整 ---
        ws.row_dimensions[row].height = new_height * 0.75
        ws.column_dimensions["B"].width = new_width / 7

        row += 1

    # === プルダウン設定 ===
    parts_list = ["上部工", "床版", "主桁", "支承", "その他"]
    damage_list = ["ひび割れ", "剥離・鉄筋露出", "漏水・遊離石灰", "その他"]
    severity_list = ["部分的・軽度", "部分的・中程度", "全体的・重度"]

    dv_parts = DataValidation(type="list", formula1='"{}"'.format(",".join(parts_list)), allow_blank=True)
    dv_damage = DataValidation(type="list", formula1='"{}"'.format(",".join(damage_list)), allow_blank=True)
    dv_severity = DataValidation(type="list", formula1='"{}"'.format(",".join(severity_list)), allow_blank=True)

    ws.add_data_validation(dv_parts)
    ws.add_data_validation(dv_damage)
    ws.add_data_validation(dv_severity)

    # 適用範囲（例：100行分）
    dv_parts.add("D2:D100")
    dv_damage.add("E2:E100")
    dv_severity.add("F2:F100")

    # ---- 保存 & ダウンロード ----
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.download_button(
        "Excel写真台帳（プルダウン＋自動調整）をダウンロード",
        data=excel_buffer,
        file_name="photo_ledger_adjusted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
